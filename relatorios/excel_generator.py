from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import pandas as pd

def gerar_excel_localizacao(df, filename, agrupar=True, exibir_unitario=True, exibir_total=True):
    """
    Gera Excel do relatório de localização
    
    Args:
        df: DataFrame com os dados
        filename: Nome do arquivo a ser salvo
        agrupar: Se True, agrupa por prateleira
    """
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Localização"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="000080")
    subtitle_font = Font(size=10)
    group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    group_font = Font(bold=True, size=11, color="000080")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    last_col_idx = 6 + (1 if exibir_unitario else 0) + (1 if exibir_total else 0)
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(last_col_idx)
    
    row = 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    cell = ws[f'A{row}']
    cell.value = "OFICINA AUTO PEÇAS"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    cell = ws[f'A{row}']
    cell.value = "POSIÇÃO DE ESTOQUE"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    cell = ws[f'A{row}']
    cell.value = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 2  # Espaço
    
    if agrupar:
        # Agrupar por prateleira
        df['GRUPO'] = df['PROD_LOCALIZACAOPECA'].apply(lambda x: str(x).split('-')[0] if pd.notna(x) else 'SEM LOC')
        
        total_geral = 0
        
        for grupo in sorted(df['GRUPO'].unique()):
            df_grupo = df[df['GRUPO'] == grupo]
            
            # Título do grupo
            ws.merge_cells(f'A{row}:{last_col_letter}{row}')
            cell = ws[f'A{row}']
            cell.value = f"GRUPO: {grupo}"
            cell.font = group_font
            cell.fill = group_fill
            row += 1
            
            # Cabeçalho da tabela
            headers = ['Código', 'Descrição', 'Marca', 'Cód. Fabricante', 'Localização', 'Estoque']
            if exibir_unitario: headers.append('Vl.Unit.')
            if exibir_total: headers.append('Vl.Total')
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            row += 1
            
            # Dados
            subtotal = 0
            for idx, produto in df_grupo.iterrows():
                ws.cell(row=row, column=1, value=produto['PROD_CODIGO']).border = border
                ws.cell(row=row, column=2, value=produto['PROD_DESCRICAOPRODUTO']).border = border
                ws.cell(row=row, column=3, value=produto.get('PROD_MARCA', '-') or '-').border = border
                ws.cell(row=row, column=4, value=produto.get('PROD_CODIGOFABRICANTE', '-') or '-').border = border
                ws.cell(row=row, column=5, value=produto.get('PROD_LOCALIZACAOPECA', '-') or '-').border = border
                
                estoque = produto.get('PROD_QTDEESTOQUEFISICO', 0) or 0
                vl_unit = produto.get('PROD_PRECOAVISTA', 0) or 0
                vl_total = produto.get('VALOR_TOTAL', 0) or 0
                
                cell = ws.cell(row=row, column=6, value=estoque)
                if estoque == int(estoque):
                    cell.number_format = '0'
                else:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                
                col_idx = 7
                if exibir_unitario:
                    cell = ws.cell(row=row, column=col_idx, value=vl_unit)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                    col_idx += 1
                
                if exibir_total:
                    cell = ws.cell(row=row, column=col_idx, value=vl_total)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                
                subtotal += vl_total
                row += 1
            
            # Subtotal
            if exibir_total:
                last_col = 6 + (1 if exibir_unitario else 0) + 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col-1)
                cell = ws.cell(row=row, column=1)
                cell.value = "Subtotal:"
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                
                cell = ws.cell(row=row, column=last_col, value=subtotal)
                cell.number_format = '#,##0.00'
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                row += 1
            
            total_geral += subtotal
            row += 2  # Espaço entre grupos
        
        # Total geral
        if exibir_total:
            last_col = 6 + (1 if exibir_unitario else 0) + 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col-1)
            cell = ws.cell(row=row, column=1)
            cell.value = "VALOR TOTAL:"
            cell.font = Font(bold=True, size=12, color="000080")
            cell.alignment = Alignment(horizontal='right')
            
            cell = ws.cell(row=row, column=last_col, value=total_geral)
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, size=12, color="000080")
            cell.alignment = Alignment(horizontal='right')
    
    else:
        # Sem agrupamento
        # Cabeçalho da tabela
        headers = ['Código', 'Descrição', 'Marca', 'Cód. Fabricante', 'Localização', 'Estoque']
        if exibir_unitario: headers.append('Vl.Unit.')
        if exibir_total: headers.append('Vl.Total')
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row += 1
        
        # Dados
        total_geral = 0
        for idx, produto in df.iterrows():
            ws.cell(row=row, column=1, value=produto['PROD_CODIGO']).border = border
            ws.cell(row=row, column=2, value=produto['PROD_DESCRICAOPRODUTO']).border = border
            ws.cell(row=row, column=3, value=produto.get('PROD_MARCA', '-') or '-').border = border
            ws.cell(row=row, column=4, value=produto.get('PROD_CODIGOFABRICANTE', '-') or '-').border = border
            ws.cell(row=row, column=5, value=produto.get('PROD_LOCALIZACAOPECA', '-') or '-').border = border
            
            estoque = produto.get('PROD_QTDEESTOQUEFISICO', 0) or 0
            vl_unit = produto.get('PROD_PRECOAVISTA', 0) or 0
            vl_total = produto.get('VALOR_TOTAL', 0) or 0
            
            cell = ws.cell(row=row, column=6, value=estoque)
            if estoque == int(estoque):
                cell.number_format = '0'
            else:
                cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            col_idx = 7
            if exibir_unitario:
                cell = ws.cell(row=row, column=col_idx, value=vl_unit)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                col_idx += 1
            
            if exibir_total:
                cell = ws.cell(row=row, column=col_idx, value=vl_total)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
            
            total_geral += vl_total
            row += 1
        
        # Total
        if exibir_total:
            last_col = 6 + (1 if exibir_unitario else 0) + 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col-1)
            cell = ws.cell(row=row, column=1)
            cell.value = "Total:"
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            cell = ws.cell(row=row, column=last_col, value=total_geral)
            cell.number_format = '#,##0.00'
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            row += 1
    
    # Ajustar largura das colunas de forma estável
    # Base 6: Cod, Desc, Marca, Ref, Local, Estoque
    larguras_fixas = {1: 10, 3: 18, 4: 15, 5: 22, 6: 12} 
    
    # Calcular largura da descrição
    total_cols = 6 + (1 if exibir_unitario else 0) + (1 if exibir_total else 0)
    desc_width = 45 + (8 if not exibir_unitario else 0) + (8 if not exibir_total else 0)
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = desc_width
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 15
    
    col_idx = 7
    if exibir_unitario:
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1
    if exibir_total:
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # Salvar
    wb.save(filename)
